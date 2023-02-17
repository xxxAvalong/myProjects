import openpyxl
import pathlib
import os
import googletrans
import shutil
from datetime import datetime
import warnings

warnings.filterwarnings('ignore',
                        message="DrawingML support is incomplete and limited to charts and images only. Shapes and drawings will be lost.",
                        category=UserWarning, module='openpyxl')
warnings.catch_warnings()

script_path_gl = pathlib.Path().resolve()

SLOVAK_ALPHABET = ['A', ' Á', ' Ä', ' B', ' C', ' Č', ' D', ' Ď', ' DZ', ' DŽ', ' E',
                   ' É', ' F', ' G', ' H', ' CH', ' I', ' Í', ' J', ' K', ' L', ' Ĺ',
                   ' Ľ', ' M', ' N', ' Ň', ' O', ' Ó', ' Ô', ' P', ' Q', ' R', ' Ŕ', ' S',
                   ' Š', ' T', ' Ť', ' U', ' Ú', ' V', ' W', ' X', ' Y', ' Ý', ' Z', ' Ž']

# global total_translated_count


def logger(txt):
    with open("logg.txt", "a", encoding="utf-8") as logg:
        logg.write(txt+f"\t{datetime.now()}")


def logg_cleaner():
    with open("logg.txt", "w", encoding="utf-8") as logg:
        logg.write("")


def is_svk_str( cell):
    cell = str(cell)
    if cell:
        for s in str(cell).strip().upper():
            if s in SLOVAK_ALPHABET:
                return True
        return False
    return False


def is_marget_cell(ws, r, c):
    # from openpyxl_image_loader import SheetImageLoader
    c_symbol = column_symbol(c)

    """image_loader = SheetImageLoader(ws)
    if image_loader.image_in(ws[f'{str(c_symbol) + str(r)}']):
        print(f"image in: {str(c_symbol) + str(r)}")
        return True"""
    cell = ws[f'{str(c_symbol) + str(r)}']
    if (cell.data_type == "s") and ("MergedCell" not in str(cell)):
        return False
    return True
    # return ("MergedCell" or "img") in str(ws[f'{str(c_symbol) + str(r)}'])


def is_excel(f_name):
    if "xls" in f_name and (("~" or "$") not in f_name):
        return ".xls"
    if "xlsx" in f_name and (("~" or "$") not in f_name):
        return ".xlsx"

    return False


def excels_from_path(script_path):
    excel_files = {}

    for path, subdirs, files in os.walk(script_path):
        for name in files:
            if is_excel(name):
                fformat = name.split('.')[-1]
                excel_files[str(os.path.join(path, name))] = [path, name, fformat]

    # excel_files = {"TotalFilePath":[FilePathOnly, FileName, FileFormat],....}
    return excel_files


EXCEL_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def column_symbol(num):
    import math

    alphabet_len = 26
    if num > alphabet_len:
        symbol_1 = EXCEL_COLUMNS[(math.ceil(num / alphabet_len) - 2)]
        symbol_2 = EXCEL_COLUMNS[num % alphabet_len - 1]

        letter = str(symbol_1 + symbol_2)
    else:
        letter = EXCEL_COLUMNS[num - 1]

    return letter


def excel_sheet_processing(w_list):
    logger(f"""\n\tSheet "{w_list}" is processing.. """)
    list_content = {}  #
    # {       key: value[]
    # txt_content: [[c1,r1], [c2,r2]]
    # } - referenced cells

    total_columns = w_list.max_column  # Sheet's columns
    total_rows = w_list.max_row  # Sheet's rows

    for c_num in range(1, total_columns):
        for r_num in range(1, total_rows):
            #try:
            cell_ = w_list.cell(column=c_num, row=r_num)  # Sheet.cell
            cell_value = (str(cell_.value).strip().replace("\n", "") if cell_.value else None)  # Sheet.cell.value
            if cell_value and (not is_marget_cell(w_list, r_num, c_num)):  # and is_svk_str(cell_value):
                try:
                    if cell_value not in list_content.keys():
                        list_content[f"{cell_value}"] = [(c_num, r_num)]
                    else:
                        list_content[f"{cell_value}"].append((c_num, r_num))
                except Exception as ex:
                    logger(f"\n\t\tCell_processing({column_symbol(c_num)}{r_num}) error: {str(ex)}")
                    continue
                    # list_content[f"{cell_value}"].append((c_num, r_num))
            #except Exception as ex:
                #print(f"Cell_processing error: {ex}")
    logger("ok\n")
    return list_content


def excel_file_processing(file_total_path, path_only, f_name):
    try:
        wf = openpyxl.load_workbook(file_total_path)  # Create work file.
    except Exception as ex:
        logger(f"\n\tError: cant open file: {f_name}\t{str(ex)}")
        return None

    all_sheets_name = wf.sheetnames
    file_content = {}  # Content of file sheets:
    # {sheet_name:
    #   {txt_content:
    #       [
    #           [c1,r1], [c2,r2] - referenced cells
    #       ]
    #   }
    # }
    for sheet_name in all_sheets_name:
        try:
            ws = wf[sheet_name]
        except Exception as ex:
            logger(f"\n\tError: cant open Sheet: {sheet_name}\t{str(ex)}")
            continue

        sheet_content = excel_sheet_processing(ws)

        file_content[f"{sheet_name}"] = sheet_content

    return file_content


def work_file(f_total_path, f_path_only, original_file_name, fformat):
    """
    Crete new, work file - "path"/file'_ua'.xlsx
    :param f_total_path: path to original file
    :param original_file_name: original file name
    :param fformat: original file format(xls or xlsx)
    :param f_path_only: only file path, without file name
    :return: path to created file
    """
    """Якщо файл ще не перекладений, тобто:
    -Не закінчується на "_ua"
    -В назві файла не присутнє "_ua"
    """
    if os.path.isfile(f_total_path.replace(f".{fformat}", "_ua.xlsx")) or original_file_name.removesuffix(
            f".{fformat}").endswith("_ua"):  # Якшо існує такий же файл в форматі(.xlsx),
        # але з приставкою "_ua" - це скоріше всього означає шо цей файл уже переводився і скоріше всього цью прогою
        return False
    else:
        new_file_name = (f_total_path.split(f'.{fformat}')[0] + "_ua" + f'.xlsx')  # Оскільки xls файли не підтримуються
        #   то в будь-якому випадку новостворений фафл буде форрмату: .xlsx
        shutil.copy(f"{f_total_path}", f"{new_file_name}")

        logger(f"\n\nFile: {original_file_name}"
               f"\nFrom: {f_path_only}"
               f"\n...is processing")

        return new_file_name


def main():
    #global total_translated_count

    logg_cleaner()
    files = excels_from_path(script_path_gl)
    # excel_files.items() = {"TotalFilePath":[FilePathOnly, FileName, FileFormat],....}

    files_count = len(files.keys())
    logger(f"\n{files_count} .xls files was found")

    for total_file_path, [f_path, f_name, f_format] in files.items():  # total_file_path = f_path, f_name + f_format

        new_work_file = work_file(f_total_path=total_file_path, f_path_only=f_path, original_file_name=f_name,
                                  fformat=f_format)

        if new_work_file:  # Якшо цільовиф файл не "_ua"
            logger(f"\n\tCreate translated file: {new_work_file}")
        else:  # Існує  файл "_ua"
            logger(
                f"""\nFile {total_file_path} is already translated. Or remove "{f_name}_ua.xlsx" file in "{f_path}". """)
            continue

        file_content = excel_file_processing(new_work_file, f_path, f_name)  # {sheet_name:{txt_content:[[c1,r1], [c2,r2] - referenced cells]}}

        translated_content = translator(file_content)  # { (ColumnRow):translated_text,... }

        f_sheets = file_content.keys()
        rewriter(file=new_work_file, f_sheets=f_sheets, content_dict=translated_content)

    # logger(f"\n\n{total_translated_count} files was translated")


def translator(file_content_dict):
    # print(f"translator: {file_content_dict}")
    translated_content = {}  # {ColumnRow):translated_text}
    for sheet, content in file_content_dict.items():
        for org_txt in content.keys():  # Оригінальний текс

            trlsted_txt = translate2(org_txt)
            for coords in content[org_txt]:  # Cell's coordinates

                x_coord = coords[0]
                column = column_symbol(x_coord)
                row = coords[1]

                cell = f"{column}{row}"
                translated_content[cell] = trlsted_txt

    return translated_content


"""def add_to_txt(file_name, sheet_name, row, column, content):
    # content=str(content)
    content_new = translate2(content)
    # print(content_new)

    if ".txt" not in file_name:

        if ".xlsx" in file_name:
            sym = ".xlsx"
        elif ".xls" in file_name:
            sym = ".xlsx"
        file_name = file_name.replace(sym, ".txt")

    with open(f"{file_name}", "a", encoding="utf-8") as txt:
        txt.write(f"\n{sheet_name}.({row},{column}):= {content} | {content_new}")"""
"""def translate(txt):
    import translate

    tt = translate.Translator(to_lang="uk", from_lang="sk")
    aa = tt.translate(txt)

    return aa"""


def rewriter(file, f_sheets, content_dict):
    # print(content_dict)
    try:
        wf = openpyxl.load_workbook(file)
    except Exception as ex:
        print(f"\nCant open _ua file. Error: \n{ex}")
        return None

    for sheet_name in f_sheets:
        try:
            ws = wf[sheet_name]
        except Exception as ex:
            print(f"\nCant open {sheet_name} in _ua file. Error: \n{ex}")
            return None

        for cell_coord, txt in content_dict.items():
            try:
                ws[cell_coord].value = txt
                # print(txt)

            except Exception as ex:
                logger(f"\t\tОшибка rewriter: {sheet_name}.{cell_coord} ")
                logger(str(ex).strip())
                continue
    wf.save(file)

    # global total_translated_count
    # total_translated_count += 1


def translate2(txt):
    # return 'rfoenrfvnsd'
    try:
        translator = googletrans.Translator()
        result = translator.translate(txt, dest="uk", src="sk").text
    except Exception as ex:
        logger(f"""\t\terror translate:"{txt}"\n {str(ex)}\nMaybe "pip install googletrans==3.1.0a0" can help """)

        return txt
    return result


if __name__ == "__main__":
    main()

